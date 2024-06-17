from openpyxl import load_workbook

def updateCell(ws, row, col, value):
    ws.cell(row=row, column=col, value=value)

def readRow(ws, row):
    return ws[row]

def readCol(ws, col):
    return [ws.cell(row=i, column=col).value for i in range(1, ws.max_row+1)]

def readCell(ws, row, col):
    return ws.cell(row=row, column=col).value

def insertRow(ws, row, values):
    ws.insert_rows(row)
    for i, value in enumerate(values):
        ws.cell(row=row, column=i+1, value=value)

def getSheetNames(wb):
    return wb.sheetnames

def getWorksheet(wb, sheetName):
    return wb[sheetName]

def dumpSheet(ws):
    for row in ws.iter_rows(values_only=True):
        print(row)

# Load an existing workbook
wb = load_workbook('REGISTRATION 2024 JAN TRAINING ON DEVELOPING QUALITY ENVIRONMENT AUDIT REPORTS IN 2024.xlsx')

# Select a specific worksheet
ws = getWorksheet(wb, getSheetNames(wb)[0])

for cell in readCol(ws, 3):
    print(cell)